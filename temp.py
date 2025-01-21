from openpyxl import load_workbook
from pymongo import MongoClient

def process_and_store_in_mongodb(file_path, db_name, chunk_size=5):
    """
    Processes an Excel file in chunks and stores data into MongoDB collections by year.

    Args:
        file_path (str): Path to the Excel file.
        db_name (str): Name of the MongoDB database.
        chunk_size (int): Number of rows to process at a time.
    """
    # Connect to MongoDB
    client = MongoClient("mongodb://localhost:27017/")
    db = client[db_name]

    # Load the workbook and target the "Current" worksheet
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["Current"]

    # Extract headers from the 5th row
    headers = [cell.value for cell in sheet[5]]
    date_columns = headers[2:]  # Columns beyond the first two are dates

    # Process rows in chunks
    for start_row in range(6, sheet.max_row + 1, chunk_size):
        chunk_data = []

        # Read the chunk of rows
        for row in sheet.iter_rows(min_row=start_row, max_row=min(start_row + chunk_size - 1, sheet.max_row), values_only=True):
            activity_name = row[0]  # Assuming "Task Name" is the first column
            if not activity_name:
                continue  # Skip rows without an activity name

            # Extract dates for this activity
            for i, cell_value in enumerate(row[2:], start=2):
                if cell_value is not None:
                    date = date_columns[i - 2]
                    year = date.split("-")[0]  # Extract year from the date

                    # Add to the chunk data
                    chunk_data.append({
                        "activity": activity_name,
                        "date": date,
                        "year": year
                    })

        # Store the chunk data into MongoDB
        for entry in chunk_data:
            collection_name = f"activities_{entry['year']}"  # Collection by year
            db[collection_name].insert_one({
                "activity": entry["activity"],
                "date": entry["date"]
            })

        print(f"Processed rows {start_row} to {min(start_row + chunk_size - 1, sheet.max_row)}")

    print("Processing and storage complete.")
    client.close()

# Usage
file_path = "Activities_Demodata_template.xlsx"
process_and_store_in_mongodb(file_path, db_name="infrastructure_activities", chunk_size=5)



















import streamlit as st
from pymongo import MongoClient
from datetime import datetime

# MongoDB connection
client = MongoClient("mongodb://localhost:27017/")
db = client["infrastructure_activities"]

# Streamlit App
st.title("Add New Activity")

# Form for adding a new activity
with st.form("add_activity_form"):
    st.subheader("New Activity Form")

    # Input fields
    activity_name = st.text_input("Activity Name", placeholder="Enter the activity name")
    activity_date = st.date_input("Activity Date", help="Select the date of the activity")
    environment = st.selectbox("Environment", ["Prod", "Non-Prod"], help="Select the environment")

    # Submit button
    submitted = st.form_submit_button("Submit")

    if submitted:
        if activity_name and activity_date:
            # Determine the year from the selected date
            year = activity_date.year

            # Prepare the data to be inserted
            activity_data = {
                "activity": activity_name,
                "date": activity_date.strftime("%Y-%m-%d"),  # Convert to string format
                "year": str(year),
                "environment": environment
            }

            # Insert into the appropriate MongoDB collection
            collection_name = f"activities_{year}"
            db[collection_name].insert_one(activity_data)

            # Success message
            st.success(f"Activity '{activity_name}' added successfully for {activity_date} in {environment} environment!")
        else:
            st.error("Please fill out all fields before submitting.")

# Display existing activities for the current year (optional)
st.subheader("Activities for the Current Year")
current_year = datetime.now().year
collection_name = f"activities_{current_year}"
activities = list(db[collection_name].find())

if activities:
    for activity in activities:
        st.write(f"- **{activity['activity']}** on {activity['date']} ({activity['environment']})")
else:
    st.info("No activities found for the current year.")
